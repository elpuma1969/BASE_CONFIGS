from netmiko import ConnectHandler
from openpyxl import Workbook
import re


def is_valid_ip(ip_addr):
    """
    This function checks if a provided string is a valid IPv4 address.

    Args:
        ip_addr (str): The string to be validated as an IP address.

    Returns:
        bool: True if the string is a valid IPv4 address, False otherwise.
    """
    pattern = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")
    return pattern.match(ip_addr) is not None


def check_device(host_ip):
    """
    This function connects to a network device and gathers information about its interfaces.

    Args:
        host_ip (str): The IP address of the network device.

    Returns:
        tuple: A tuple containing the device hostname and a list of interfaces in the "notconnect" state.
    """
    device = {
        "device_type": "cisco_ios",
        "host": host_ip,
        "username": "puma",
        "password": "cisco",
    }

    with ConnectHandler(**device) as net_connect:
        try:
            hostname = net_connect.send_command("show running-config | include hostname").split(" ")[-1].strip()
            output = net_connect.send_command("show interface status")
            fa_notconnect = []

            for line in output.splitlines():
                if "Fa" in line:
                    interface, state = line.split()[0:2]
                    if state == "notconnect":
                        fa_notconnect.append(interface)

            return hostname, fa_notconnect
        except Exception as e:
            print(f"Error connecting to device: {host_ip} - {e}")
            return None, None


# Create a new Excel workbook
wb = Workbook()

# Read IP addresses from the file
try:
    with open("devices_file", "r") as file:
        ip_addresses = file.read().splitlines()
except FileNotFoundError:
    print("Error: File 'devices_file' not found.")
    exit()

# Check each device and add data to the workbook
for ip in ip_addresses:
    if is_valid_ip(ip):
        print(f"Checking IP: {ip}")
        hostname, not_connect_interfaces = check_device(ip)

        if hostname and not_connect_interfaces:
            ws = wb.create_sheet(title=hostname)
            ws.append(["Interface", "Status"])

            for interface in not_connect_interfaces:
                ws.append([interface, "notconnect"])

            print(f"Device '{hostname}' added to the workbook.")
        elif hostname:
            ws = wb.create_sheet(title=hostname)
            ws.append(["Interface", "Status"])
            ws.append(["All interfaces", "operational"])
            print(f"All interfaces on device '{hostname}' are operational.")
    else:
        print(f"Invalid IP address: {ip}")

# Remove default sheet
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# Save the workbook
try:
    wb.save("network_devices_status.xlsx")
    print("Workbook saved successfully.")
except Exception as e:
    print(f"Error saving workbook: {e}")
