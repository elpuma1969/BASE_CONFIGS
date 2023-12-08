from getpass import getpass
from netmiko import ConnectHandler
from datetime import datetime
import os
from openpyxl import load_workbook


def read_excel(file_name):
    # Reads an Excel file and returns IPs from column C
    workbook = load_workbook(file_name)
    sheet = workbook.active
    devices = []
    for row in range(2, sheet.max_row + 1):
        ip = sheet[f'C{row}'].value
        if ip is not None:
            devices.append(ip)
    return devices


def read_file(file_name):
    # Reads a file and returns its content line by line
    with open(file_name) as f:
        content = f.read().splitlines()
    return content


def connect_to_device(device_ip, device_type, username, password, port=22):
    # Establishes an SSH connection to a network device
    device = {
        'device_type': device_type,
        'ip': device_ip,
        'username': username,
        'password': password,
        'port': port,
    }
    return ConnectHandler(**device)


def execute_commands(connection, commands):
    # Executes a list of commands on the connected device
    connection.enable()
    return connection.send_config_set(commands)


def save_output(output, hostname):
    # Saves command outputs to a file with a timestamp
    now = datetime.now()
    directory = 'backups'
    if not os.path.exists(directory):
        os.makedirs(directory)
    filename = f'{directory}/{hostname}_{now.year}-{now.month}-{now.day}_update.txt'
    with open(filename, 'w') as final:
        final.write(output)
        print(f'Backup of {hostname} completed successfully')
        print('#' * 30)


def main():
    # Main function to orchestrate the script's workflow
    username = input("Enter your SSH username:")
    password = getpass()

    commands_list = read_file('command_file')
    devices_list = read_excel('TEST_LAB.xlsx')  # Replace with your Excel file name

    for device_ip in devices_list:
        print("Connecting to device: " + device_ip)

        try:
            connection = connect_to_device(device_ip, 'cisco_ios', username, password)
            print("Entering the enable mode...")

            output = execute_commands(connection, commands_list)

            prompt = connection.find_prompt()
            hostname = prompt[0:-1]

            save_output(output, hostname)

            print("Closing connection")
            connection.disconnect()
        except Exception as e:
            print(f"Failed to connect or execute commands on {device_ip}: {e}")


if __name__ == "__main__":
    main()
